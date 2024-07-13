import argparse
import re

import pandas as pd
import numpy as np
from openpyxl import load_workbook


def create_parser():
    parser = argparse.ArgumentParser(description="""Process an XLSX file that
    contains the official Global Industry Classification Standard (GICS)
    released by MSCI, transforming and normalizing the data, and outputting the
    results as a CSV file.""")
    parser.add_argument("filename", type=str,
                        help="The path to the XLSX file to process.")
    parser.add_argument("--sheet", type=str, default=0,
                        help="Sheet in the XLSX to use. Default: first sheet.")
    parser.add_argument("--output", type=str, default="output.csv",
                        help="Output CSV file name. Default: output.csv.")
    parser.add_argument("--drop-discontinued", type=bool, default=False,
                        help="Remove discontinued classifications. Default: False.")
    return parser


def get_codes_to_remove(filename, sheet):
    wb = load_workbook(filename=filename, data_only=True)
    if sheet >= len(wb.sheetnames):
        raise IndexError(
            f"Sheet index {sheet} is out of range. Available sheets: {wb.sheetnames}")

    ws = wb[wb.sheetnames[sheet]]
    pattern = r'\(.*?\bdiscontinued\b.*?\)'
    codes = []

    # The code for a classification is added to the `code` array if a cell is
    # found to contain the strikethrough formatting or the word "discontinued"
    # is found enclosed within parentheses in a cell.
    for row in ws.iter_rows():
        for cell in row:
            if not cell.font.strikethrough and not re.search(
                    pattern, str(cell.value), re.IGNORECASE):
                continue

            code = row[6].value
            if code is not None:
                codes.append(code)

            break

    wb.close()
    return codes


def normalize_merged_cols(cols, subject, code_col, name_col):
    tgt_col_idx = cols.index(subject)
    cols[tgt_col_idx] = code_col
    cols[tgt_col_idx + 1] = name_col


def main(args):
    # Get a list of codes to remove (done further down below), if applicable.
    codes_to_remove = get_codes_to_remove(
        args.filename, args.sheet) if args.drop_discontinued else []

    df = pd.read_excel(args.filename, sheet_name=args.sheet, header=None)

    # Identify header row
    header_row = df[df.apply(
        lambda y: y.map(lambda x: "Sector" in str(x))).any(axis=1)].index[0]

    # Remove rows above the dataframe's header
    df.columns = df.loc[header_row]
    df = df.loc[header_row + 1:]
    df.reset_index(drop=True, inplace=True)

    # Give header columns appropriate names and add a column for the industry
    # description.
    columns_list = df.columns.tolist()
    normalize_merged_cols(columns_list, 'Sector', 'sector_code', 'sector_name')
    normalize_merged_cols(columns_list, 'Industry Group',
                          'industry_group_code', 'industry_group_name')
    normalize_merged_cols(columns_list, 'Industry',
                          'industry_code', 'industry_name')
    normalize_merged_cols(columns_list, 'Sub-Industry',
                          'sub_industry_code', 'sub_industry_name')
    df.columns = columns_list

    # Force string cells that are blank to NA
    df = df.apply(lambda col: col.map(lambda x: np.nan if isinstance(
        x, str) and re.match(r'^\s*$', x) else x))

    # Copy industry descriptions from subsequent rows into the
    # 'industry_description' column.  This process identifies rows where the
    # 'sub_industry_code' is missing (indicating description rows) and
    # concatenates these descriptions into the 'industry_description' field of
    # the preceding row that contains a 'sub_industry_code'. After
    # concatenation, these description rows will be skipped in the iteration to
    # ensure they are not processed again.
    df['industry_description'] = ''
    i = 0
    while i < len(df):
        if pd.isna(df.at[i, 'sub_industry_code']):
            j = i
            while j < len(df) and pd.isna(df.at[j, 'sub_industry_code']):
                df.at[i-1, 'industry_description'] += " " + \
                    str(df.at[j, 'sub_industry_name']).strip()
                j += 1
            i = j
        else:
            i += 1

    # Remove rows that do not contain a sub_industry_code
    df = df[df['sub_industry_code'].notna()].reset_index(drop=True)

    # Remove rows that refer to classifications that have been discontinued
    if args.drop_discontinued and len(codes_to_remove) > 0:
        df = df[~df['sub_industry_code'].isin(codes_to_remove)]

    # Make it so all NA columns inherit the content of the preceding row
    df['sector_code'] = df['sector_code'].ffill().astype(int)
    df['sector_name'] = df['sector_name'].ffill()
    df['industry_group_code'] = df['industry_group_code'].ffill().astype(int)
    df['industry_group_name'] = df['industry_group_name'].ffill()
    df['industry_code'] = df['industry_code'].ffill().astype(int)
    df['industry_name'] = df['industry_name'].ffill()
    df['sub_industry_code'] = df['sub_industry_code'].ffill().astype(int)
    df['sub_industry_name'] = df['sub_industry_name'].ffill()

    # Fixing more oddities:
    # - Collapse multiple spaces into one
    df = df.replace(r'\s\s+', ' ', regex=True)
    # - Remove in-column annotations
    df = df.replace(r'\([^)]*\)', '', regex=True)
    # - Strip wrapped quotes
    df = df.apply(lambda y: y.map(
        lambda x: x[1:-1] if isinstance(x, str) and re.match(r'^".*"$', x) else x))

    # Fix whitespace
    df = df.apply(lambda y: y.map(
        lambda x: x.strip() if isinstance(x, str) else x))
    df = df.apply(lambda y: y.map(lambda x: x.replace(
        '\n', '') if isinstance(x, str) else x))

    # Done!  Persist to CSV file.
    df.to_csv(args.output, index=False)


if __name__ == '__main__':
    parser = create_parser()
    main(parser.parse_args())
